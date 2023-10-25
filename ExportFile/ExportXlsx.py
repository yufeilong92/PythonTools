#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2023/7/15 16:12
# @Author  : backpacker
# @File    : ExportXlsx.py
# @Description : $导出XLsx
import os.path
import time
from tkinter import *
from tkinter import  filedialog

from tkinter import  messagebox

from openpyxl import Workbook,load_workbook

from Base.MainQuit import MainQuit


class ExportXlsx(MainQuit):
    def adbSelectFilePath(self,selectPathTv,isSave):
        selectFilePath = filedialog.askdirectory()
        if selectFilePath is None or selectFilePath=="":
            if isSave:
                selectFilePath="保存的目录"
            else:
                selectFilePath="导出的目录"
        selectPathTv.config(text=selectFilePath)
        pass
    def showDialog(self,mainRoot):

        root = Tk()
        root.config(bg="light gray")
        root.geometry("640x540")#设置窗口大小
        mFont = 10
        mSticky = W + E + S + N
        mWraplength = 300,
        #=================================
        Label(master=root,text="地址：",font=mFont).grid(row=0,column=0,sticky=mSticky,pady=4)
        selectPathTv = Label(master=root, text="导出的目录",wraplength=mWraplength, font=mFont)
        selectPathTv.grid(row=0, column=1, sticky=W + E + S + N,pady=4)
        Button(master=root,text="选择",font=mFont,command=lambda :self.adbSelectFilePath(selectPathTv,False)).grid(row=0,column=2,sticky=W+E+S+N,pady=4)
        #标题权重
        childrenLenght = root.children.__len__()
        for i in range(childrenLenght):
            root.columnconfigure(i,weight=1)

        Label(master=root,text="地址：",font=mFont).grid(row=1,column=0,sticky=E+W+S+N,pady=4)
        saveFilePathTv = Label(master=root, text="保存的目录", wraplength=mWraplength,font=mFont)
        saveFilePathTv.grid(row=1, column=1, sticky=W + E + S + N,pady=4)
        Button(master=root,font=mFont,text="选择",command=lambda :self.adbSelectFilePath(saveFilePathTv,True)).grid(row=1,column=2,sticky=W+S+N+E,pady=4)


        Label(master=root,text="文件名称",font=mFont).grid(row=2,column=0,sticky=W+E+S+N,pady=4)

        saveName = Entry(master=root, font=mFont)
        saveName.grid(row=2,column=1,sticky=N+W+S+E,pady=4)

        Label(master=root,text=".xlsx",font=mFont).grid(row=2,column=2,sticky=E+W+S+N,pady=4)


        Button(master=root,text="保存",font=mFont,command=lambda :self.adbSave(selectPathTv,saveFilePathTv,saveName)).grid(row=3,column=1,sticky=W+N+S+E,pady=8)

        super().registLisetener(root,mainRoot)
        root.mainloop()

        pass

    def adbSave(self, selectPathTv, saveFilePathTv, saveName):
        saveName = saveName.get()
        selectPath = selectPathTv.cget("text")
        savePath = saveFilePathTv.cget("text")

        print(f"saveName{saveName}")
        print(f"selectPath{selectPath}")
        print(f"savePath{savePath}")

        if selectPath is None or selectPath=="导出的目录":
            messagebox.showwarning("温馨提示","请选择导出目录")
            return
        if savePath is None or savePath =="保存的目录":
            messagebox.showwarning("温馨提示","请选择保存目录")
            return
        if saveName is None or saveName == "":
            messagebox.showwarning("温馨提示","请输入文件名称")
            return
        listdir = os.listdir(selectPath)
        if listdir.__len__() == 0:
            messagebox.showwarning("温馨提示","选择目录下没有内容")
            return

        askquestion = messagebox.askquestion("温馨提示", "是否确认保存")

        if askquestion !="yes":
            print("退出")
            return
        mSaveOverPath=savePath+"/"+saveName+".xlsx"

        wb:Workbook
        if  not os.path.exists(mSaveOverPath) :
            print(f"不存在,创建{mSaveOverPath}")
            wb=Workbook()
            #创建表名
            sh = wb.create_sheet(saveName)

            self.saveXlsx(listdir,wb,sh,1,mSaveOverPath,savePath)
            return
            # for postion in range(listdir.__len__()):
            #     sh.cell(row=postion+1,column=1).value=f"{listdir.__getitem__(postion)}"
            # wb.save(mSaveOverPath)
            # wb.close()
            # if messagebox.askquestion("温馨提示", "保存成功，是否跳转文件目录") == "yes":
            #     os.startfile(savePath)
            #     return
        print("存在")
        wb=load_workbook(mSaveOverPath)
        sheetnames = wb.sheetnames
        #判断是否存在
        isExitesSheetName=-1
        for indext in range(len(sheetnames)):
            if saveName==sheetnames[indext]:
                isExitesSheetName=indext
                break
        if isExitesSheetName==-1:#不存在，则创建
            wb.create_sheet(saveName)
            sh=wb[saveName]
            self.saveXlsx(listdir,wb,sh,1,mSaveOverPath,savePath)
            # for postion in range(listdir.__len__()):
            #     sh.cell(row=postion + 1, column=1).value = f"{listdir.__getitem__(postion)}"
            # wb.save(mSaveOverPath)
            # wb.close()
            # if messagebox.askquestion("温馨提示", "保存成功，是否跳转文件目录") == "yes":
            #     os.startfile(savePath)
            #     return
        else:
            if messagebox.askquestion(title="温馨提示", message="是否追加，否则创建") == "yes":
            #表已经存在，追加
                sh=wb[sheetnames[isExitesSheetName]]
                exitesPostion=sh.max_row+1

                self.saveXlsx(listdir,wb,sh,exitesPostion,mSaveOverPath,savePath)

                # for postion in range(listdir.__len__()):
                #     sh.cell(row=exitesPostion+postion, column=1).value = f"{listdir.__getitem__(postion)}"
                # wb.save(mSaveOverPath)
                # wb.close()
                # if messagebox.askquestion("温馨提示", "追加成功，是否跳转文件目录") == "yes":
                #     os.startfile(savePath)
                #     return
            else:
                createTime = time.strftime("%Y-%M-%d %H-%M-%S", time.localtime())
                saveNewName=saveName+f"{createTime}"
                print(f"savenewanem{saveNewName}")
                wb.create_sheet(saveNewName)
                sh=wb[saveNewName]

                self.saveXlsx(listdir,wb,sh,1,mSaveOverPath,savePath)

                # for postion in range(listdir.__len__()):
                #     sh.cell(row=postion + 1, column=1).value = f"{listdir.__getitem__(postion)}"
                # wb.save(mSaveOverPath)
                # wb.close()
                # if messagebox.askquestion("温馨提示", "保存成功，是否跳转文件目录") == "yes":
                #     os.startfile(savePath)
                #     return
        pass
    #  exitesPostion 表格起始位置，
    #  mSaveOverPath 保存位置。
    #  savePath保存目录
    def saveXlsx(self,data:list[str],wb:Workbook,sh,exitesPostion:int,mSaveOverPath,savePath):
        try:
            for postion in range(data.__len__()):
              sh.cell(row=exitesPostion + postion, column=1).value = f"{data.__getitem__(postion)}"

            wb.save(mSaveOverPath)
            wb.close()
            if messagebox.askquestion("温馨提示", "保存成功，是否跳转文件目录") == "yes":
                os.startfile(savePath)
                return
        except:
            messagebox.showerror("温馨提示",f"要保存的文件{mSaveOverPath}正在打开，或者异常，请关闭重试")
        pass