#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2023/7/16 15:24
# @Author  : backpacker
# @File    : QueryRep.py
# @Description : $查询重复
import os
import time
from tkinter import *
from tkinter import filedialog, messagebox, Tk
from tkinter.ttk import Combobox

from Base.FileDialogType import FileDialogType
from Base.MainQuit import MainQuit

from openpyxl import Workbook, load_workbook


class QueryRep(MainQuit):
    def initConstant(self, str: StringVar, strOne: StringVar):
        str.set("")
        strOne.set("")
        pass

    def selectDialogMater(self, tv: Label, type: FileDialogType, oldHint: str, newHint: str, isNew: bool,
                          str: StringVar, strOne: StringVar):
        path = None
        if type == FileDialogType.DIRECTORY:
            dialog = filedialog.askdirectory()
            if dialog is not None:
                path = dialog
        elif type == FileDialogType.DOCUMENT:  # 文件
            fileType = [("Xlsx", ".xlsx")]
            dialog = filedialog.askopenfile(filetypes=fileType)
            if dialog is not None:
                path = dialog.name

        if path is None or path == "":
            path = newHint if isNew else oldHint

        oldPath = tv.cget("text")
        tv.config(text=path)
        if oldPath != path:
            self.initConstant(str, strOne)
        pass

    def showDialog(self, mainRoot):

        root = Tk()
        root.title("查重工具")
        root.config(bg="light gray")
        root.geometry("640x540")  # 设置窗口大小
        mFont = 10
        mSticky = W + E + S + N
        mWraplength = 300,
        mRlief = "groove"

        rowIndex = 0
        mStrVar = StringVar(master=root)
        mStrVarTwo = StringVar(master=root)

        Label(master=root, font=mFont, text="地址：", relief=mRlief).grid(row=rowIndex, column=0, sticky=mSticky, pady=4)
        selectFileTv = Label(master=root, font=mFont, wraplength=mWraplength, text="选择文件", relief="groove")
        selectFileTv.grid(row=rowIndex, column=1, sticky=mSticky, pady=4)
        Button(master=root, font=mFont, text="选择", command=lambda: self.
               selectDialogMater(selectFileTv, FileDialogType.DOCUMENT, "选择文件", "", False, mStrVar,
                                 mStrVarTwo)).grid(row=rowIndex, column=2, sticky=mSticky, pady=4)

        childrenLenght = root.children.__len__()
        for postion in range(childrenLenght):
            root.columnconfigure(postion, weight=1)
        rowIndex += 1
        Label(master=root, text="表名：", font=mFont, relief=mRlief).grid(row=rowIndex, column=0, sticky=mSticky, pady=4)
        combobox = Combobox(master=root, font=mFont, textvariable=mStrVar,
                            postcommand=lambda: self.adbComboChange(mStrVarTwo))
        combobox.grid(row=rowIndex, column=1, sticky=mSticky, pady=4)
        Button(master=root, text="获取表名", font=mFont, relief=mRlief,
               command=lambda: self.adbGetSheetName(selectFileTv, combobox, mStrVar, mStrVarTwo)).grid(row=rowIndex,
                                                                                                       column=2,
                                                                                                       sticky=mSticky,
                                                                                                       pady=4)

        rowIndex += 1
        Label(master=root, text="第一行数据：", font=mFont, relief=mRlief).grid(row=rowIndex, column=0, sticky=mSticky,
                                                                               pady=4)
        comboboxFirst = Combobox(master=root, font=mFont, textvariable=mStrVarTwo)
        comboboxFirst.grid(row=rowIndex, column=1, sticky=mSticky, pady=4)
        Button(master=root, text="获取数据", font=mFont, relief=mRlief,
               command=lambda: self.adbGetSheetData(selectFileTv, mStrVar, mStrVarTwo, comboboxFirst)).grid(
            row=rowIndex, column=2, sticky=mSticky, pady=4)

        rowIndex += 1
        Label(master=root, text="地址：", font=mFont, relief=mRlief).grid(row=rowIndex, column=0, sticky=mSticky, pady=4)
        saveFileTv = Label(master=root, text="保存的地址", font=mFont, wraplength=mWraplength, relief=mRlief)
        saveFileTv.grid(row=rowIndex, column=1, sticky=mSticky, pady=4)
        Button(master=root, text="选择", font=mFont, relief=mRlief,
               command=lambda: self.selectFileDialog(saveFileTv, FileDialogType.DIRECTORY, "保存的地址", "",
                                                     False)).grid(row=rowIndex, column=2, sticky=mSticky, pady=4)

        rowIndex += 1
        Label(master=root, text="文件名称：", font=mFont, relief=mRlief).grid(row=rowIndex, column=0, sticky=mSticky,
                                                                             pady=4)
        entry = Entry(master=root, font=mFont)
        entry.grid(row=rowIndex, column=1, sticky=mSticky, pady=4)
        saveTypeTv = Label(master=root, text="", font=mFont, relief=mRlief)
        saveTypeTv.grid(row=rowIndex, column=2, sticky=mSticky, pady=4)

        rowIndex += 1
        mRepVar = IntVar(master=root)
        mRepVar.set(0)
        Label(master=root, text="保存", font=mFont, relief=mRlief).grid(row=rowIndex, column=0, sticky=mSticky, pady=4)
        Radiobutton(master=root, text="去重的数据", value=0, variable=mRepVar, font=mFont, relief=mRlief,
                    command=lambda: self.adbRadiobutton(mRepVar, saveTypeTv)).grid(row=rowIndex, column=1,
                                                                                   sticky=mSticky, pady=4)
        Radiobutton(master=root, text="重复的数据", value=1, variable=mRepVar, font=mFont, relief=mRlief,
                    command=lambda: self.adbRadiobutton(mRepVar, saveTypeTv)).grid(row=rowIndex, column=2,
                                                                                   sticky=mSticky, pady=4)

        rowIndex += 1
        mIntVar = IntVar(master=root)
        mIntVar.set(0)
        Radiobutton(master=root, text="查重保存新的表", value=0, variable=mIntVar, font=mFont, relief=mRlief,
                    command=lambda: self.adbRadiobutton(mIntVar, saveTypeTv)).grid(row=rowIndex, column=0,
                                                                                   sticky=mSticky, pady=4)
        Radiobutton(master=root, text="查重保存新的xlsx", value=1, variable=mIntVar, font=mFont, relief=mRlief,
                    command=lambda: self.adbRadiobutton(mIntVar, saveTypeTv)).grid(row=rowIndex, column=1,
                                                                                   sticky=mSticky, pady=4)
        Radiobutton(master=root, text="查重保存新的TXT", value=2, variable=mIntVar, font=mFont, relief=mRlief,
                    command=lambda: self.adbRadiobutton(mIntVar, saveTypeTv)).grid(row=rowIndex, column=2,
                                                                                   sticky=mSticky, pady=4)

        rowIndex += 1
        Button(master=root, text="确认", font=mFont,
               command=lambda: self.adbSave(mIntVar.get(), selectFileTv, saveFileTv, entry,
                                            mRepVar, combobox, comboboxFirst)).grid(row=rowIndex, column=1,
                                                                                    sticky=mSticky, pady=5)

        super().registLisetener(root, mainRoot)
        mainloop()

        pass

    def QueryWindow(self, root: Tk, mainRoot: Tk):
        root.destroy()
        mainRoot.deiconify()

    pass

    def adbSave(self, param: int, selectFileTv: Label, saveFileTv, saveName: Entry, mRepVAR: IntVar,
                sheelCombox: Combobox, rowCombobx: Combobox):
        mSheelLenght = sheelCombox.current()
        mRowLength = rowCombobx.current()
        repVarType = mRepVAR.get()
        print(f"starvar{mSheelLenght}")
        # 选择xlsx
        selectPath = selectFileTv.cget("text")
        savePath = saveFileTv.cget("text")
        saveName = saveName.get()
        if selectPath is None or selectPath == "选择文件":
            messagebox.showwarning("温馨提示", "请选择查重的文件")
            return

        if mSheelLenght is None or mSheelLenght == -1:
            messagebox.showwarning("温馨提示", "请选择表名")
            return

        if mRowLength is None or mRowLength == -1:
            messagebox.showwarning("温馨提示", "请选择某列")
            return

        if saveName is None or saveName == "":
            messagebox.showwarning("温馨提示", "请输入名称")
            return

        wb: Workbook = None
        wb = load_workbook(selectPath)
        sheetnames = wb.sheetnames
        lenght_ = sheetnames[mSheelLenght]
        sh = wb[lenght_]
        rows = sh.rows
        oldList = []
        # oldList=[ (children.value for children in parent)  for parent in rows]

        for item in rows:
            if item is not None:
                children = item[mRowLength]
                oldList.append(children.value)
        print(f"原有数据{oldList}")
        deweightlist = []  # 去重的数据
        repetitionList = []  # 重复数据
        if oldList.__len__() == 1:  # 如果数据为一直接复制
            deweightlist = oldList
            repetitionList = oldList
        else:
            for item in oldList:
                if item not in deweightlist:  # 不存在
                    deweightlist.append(item)
                else:  # 存在
                    if item not in repetitionList:
                        repetitionList.append(item)
        wb.close()
        if param == 0:  # 查重保存新的表
            if saveName in sheetnames:
                createTime = time.strftime("%Y-%m-%d %H-%M-%S", time.localtime())
                newName = saveName + (createTime)
            else:
                newName = saveName
            print(newName)
            self.saveDataXlsx(deweightlist, newName, repVarType, repetitionList, selectPath, False, True)
            pass
        elif param == 1:  # 去重保存新的xlsx
            if savePath is None or savePath == "保存的地址":
                messagebox.showwarning("温馨提示", "请选择保存目录")
                return
            newPath = savePath + "/" + saveName + ".xlsx"
            if os.path.exists(newPath):
                messagebox.showwarning("温馨提示", f"保存的目录{newPath}已经存在")
                return
            self.saveDataXlsx(deweightlist, saveName, repVarType, repetitionList, newPath, False, False)
            pass
        elif param == 2:  # 重复保存新的TXT
            if savePath is None or savePath == "保存的地址":
                messagebox.showwarning("温馨提示", "请选择保存目录")
                return
            newPath = savePath + "/" + saveName + ".txt"
            if os.path.exists(newPath):
                messagebox.showwarning("温馨提示", f"保存的目录{newPath}已经存在")
                return
            self.saveDataXlsx(deweightlist, saveName, repVarType, repetitionList, newPath, True, False)
            pass

        pass

    def saveDataXlsx(self, deweightlist, newName, repVarType, repetitionList, selectPath, isText: bool,
                     isSaveOld: bool):
        if isText:
            file = open(file=selectPath, mode="w", encoding="utf-8")
            if repVarType == 0:
                for item in deweightlist:
                    file.write(f"{item}\n")
                file.close()
            else:
                if repetitionList.__len__() == 0:
                    messagebox.showwarning("温馨提示", "无重复数据")
                    return
                for item in repetitionList:
                    file.write(f"{item}\n")
                file.close()
            if messagebox.askquestion("温馨提示", "保存成功，是否跳转文件目录") == "yes":
                s = str(selectPath)
                print(s)
                rfind = s.rfind("/")
                rfind_ = s[:rfind]
                print(rfind_)
                os.startfile(rfind_)
            return
        print(f"newName={newName}")
        wb: Workbook = None
        if isSaveOld:
            wb = load_workbook(selectPath)
        else:
            wb = Workbook()
        wb.create_sheet(newName)
        sh = wb[newName]
        if repVarType == 0:  # 去重数据
            for index in range(len(deweightlist)):
                sh.cell(row=index + 1, column=1).value = f"{deweightlist[index]}"
        else:
            if repetitionList.__len__() == 0:
                messagebox.showwarning("温馨提示", "无重复数据")
                return
            for index in range(len(repetitionList)):
                sh.cell(row=index + 1, column=1).value = f"{repetitionList[index]}"
        try:
            wb.save(selectPath)
            wb.close()
            if messagebox.askquestion("温馨提示", "保存成功，是否跳转文件目录") == "yes":
                s = str(selectPath)
                print(s)
                rfind = s.rfind("/")
                rfind_ = s[:rfind]
                print(rfind_)
                os.startfile(rfind_)
        except:
            messagebox.showwarning("温馨提示", f"保存的文件{selectPath}是否已经打开，或者异常")

    def adbRadiobutton(self, mIntVar, saveTypeTv):
        param = mIntVar.get()
        print(f"param{param}")
        if param == 0:  # 查重保存新的表
            saveTypeTv.config(text="")
            pass
        elif param == 1:  # 查重保存新的xlsx
            saveTypeTv.config(text=".xlsx")
            pass
        elif param == 2:  # 查重保存新的TXT
            saveTypeTv.config(text=".txt")
            pass

        pass

    def adbGetSheetName(self, selectFileTv, combobox: Combobox, mStrVar: StringVar, mStrTwo: StringVar):
        selectPath = selectFileTv.cget("text")

        if selectPath is None or selectPath == "选择文件":
            messagebox.showwarning("温馨提示", "请选择文件")
            return
        wb: Workbook = None
        wb = load_workbook(selectPath)
        sheetnames = wb.sheetnames
        combobox['value'] = sheetnames
        combobox.current(0)
        mStrVar.set(sheetnames[0])
        mStrTwo.set("")
        wb.close()
        pass

    def adbGetSheetData(self, selectFileTv: Label, mStrVar: StringVar, mStrVarFirst: StringVar,
                        comboboxFirst: Combobox):

        selectPath = selectFileTv.cget("text")
        if selectPath is None or selectPath == "选择文件":
            messagebox.showwarning("温馨提示", "请选择文件")
            return
        value = mStrVar.get()

        if value is None or value == "":
            messagebox.showwarning("温馨提示", "请选择表名")
            return

        wb: Workbook = None
        wb = load_workbook(selectPath)
        sh = wb[value]
        columns = sh.max_column
        # rowsMax = sh.max_row
        rows = sh.rows
        # rowValuesDebug=[]
        rowValues = []
        index = 1
        # 判断是否终止
        isConice = False
        print(f"总共几列{columns}")
        for item in rows:
            if isConice:
                break
            indexColumns = 1
            for itemPostion in item:
                # rowvalue=f"第{index}行，第{indexColumns}列，值={itemPostion.value}"
                # print(rowvalue)
                # rowValuesDebug.append(rowvalue)
                rowValues.append(itemPostion.value)
                if indexColumns == columns:
                    isConice = True
                    break
                indexColumns += 1
            index += 1
        # print(f"所有数据{rowValuesDebug}")
        # print(f"第一列数据 {rowValues}")
        if rowValues.__len__() == 0:
            messagebox.showwarning("温馨提示", "选择表中，第一行没有数据")
            return
        comboboxFirst["value"] = rowValues
        comboboxFirst.current(0)
        mStrVarFirst.set(rowValues[0])
        wb.close()
        pass
