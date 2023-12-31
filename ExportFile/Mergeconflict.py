import datetime
import os
from tkinter import Tk, filedialog, messagebox
from tkinter import *
from tkinter.ttk import Combobox

from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook

from Base.FileDialogType import FileDialogType
from Base.MainQuit import MainQuit

from tkinter.filedialog import askopenfilename


class Mergeconflict(MainQuit):

    def showDialog(self, mainRoot):
        root = Tk()
        root.title("查重工具")
        root.config(bg="light gray")
        root.geometry("700x650")  # 设置窗口大小
        mFont = 10
        mSticky = W + E + S + N
        mWraplength = 300,
        mRlief = "groove"
        rowNum = 0
        mOneSheet = StringVar(master=root)
        mOneRow = StringVar(master=root)
        mTwoSheet = StringVar(master=root)
        mTwoRow = StringVar(master=root)

        mIntVar = IntVar(master=root)
        mIntVar.set(0)
        # ==============================================================
        Label(master=root, text="文件1：", font=mFont, relief=mRlief).grid(row=rowNum, column=0, sticky=mSticky,
                                                                          pady=4)
        selectPathOneTv = Label(master=root, text="One文件目录", wraplength=mWraplength, font=mFont, relief=mRlief)
        selectPathOneTv.grid(row=rowNum, column=1, sticky=mSticky, pady=4)
        Button(master=root, text="选择", font=mFont, relief=mRlief, command=lambda: self.
               adbSelectFileOne(selectPathOneTv, mOneSheet, mOneRow, True,0)).grid(row=rowNum, column=2,
                                                                                 sticky=mSticky, pady=4)
        len__ = root.children.__len__()
        for index in range(len__):
            root.columnconfigure(index, weight=1)
        # ==============================================================
        rowNum += 1
        Label(master=root, text="文件1表名：", font=mFont, relief=mRlief).grid(row=rowNum, column=0, sticky=mSticky,
                                                                              pady=4)
        oneSheetCbb = Combobox(master=root, textvariable=mOneSheet, font=mFont,
                               postcommand=lambda: self.adbComboChange(mOneRow))
        oneSheetCbb.grid(row=rowNum, column=1, sticky=mSticky, pady=4)
        Button(master=root, text="获取", font=mFont, relief=mRlief, command=lambda: self.
               adbSelectSheeet(mOneRow, selectPathOneTv, oneSheetCbb, True)).grid(row=rowNum, column=2,
                                                                                           sticky=mSticky,
                                                                                           pady=4)

        rowNum += 1
        Label(master=root, text="文件1第一行：", font=mFont, relief=mRlief).grid(row=rowNum, column=0, sticky=mSticky,
                                                                                pady=4)
        oneRowCbb = Combobox(master=root, textvariable=mOneRow, font=mFont)
        oneRowCbb.grid(row=rowNum, column=1, sticky=mSticky, pady=4)
        Button(master=root, text="获取", font=mFont, relief=mRlief, command=lambda: self.
               adbSelectRowsName(selectPathOneTv, oneSheetCbb, oneRowCbb, True)).grid(row=rowNum, column=2,
                                                                                      sticky=mSticky, pady=4)

        rowNum += 1
        Label(master=root, text="选择方式", font=mFont, relief=mRlief).grid(row=rowNum, column=0, sticky=mSticky,
                                                                            pady=4)
        Radiobutton(master=root, text="文件2", value=0, variable=mIntVar, font=mFont, relief=mRlief).grid(row=rowNum,
                                                                                                          column=1,
                                                                                                          sticky=mSticky,
                                                                                                          pady=4)
        Radiobutton(master=root, text="新的目录", value=1, variable=mIntVar, font=mFont, relief=mRlief).grid(row=rowNum,
                                                                                                             column=2,
                                                                                                             sticky=mSticky,
                                                                                                             pady=4)
        rowNum += 1
        Label(master=root, text="文件2：", font=mFont, relief=mRlief).grid(row=rowNum, column=0, sticky=mSticky,
                                                                          pady=4)
        selectPathTwoTv = Label(master=root, text="Two文件目录", font=mFont, relief=mRlief, wraplength=mWraplength)
        selectPathTwoTv.grid(row=rowNum, column=1, sticky=mSticky, pady=4)
        Button(master=root, text="选择", font=mFont, relief=mRlief, command=lambda: self.
               adbSelectFileOne(selectPathTwoTv, mTwoSheet, mTwoRow, False, mIntVar)).grid(row=rowNum, column=2,
                                                                                           sticky=mSticky,
                                                                                           pady=4)

        rowNum += 1
        Label(master=root, text="文件2表名：", font=mFont, relief=mRlief).grid(row=rowNum, column=0, sticky=mSticky,
                                                                              pady=4)
        twoSheetCbb = Combobox(master=root, textvariable=mTwoSheet, font=mFont,
                               postcommand=lambda: self.adbComboChange(mTwoRow))
        twoSheetCbb.grid(row=rowNum, column=1, sticky=mSticky, pady=4)
        Button(master=root, text="获取", font=mFont, relief=mRlief, command=lambda: self.
               adbSelectSheeet(mTwoRow, selectPathTwoTv, twoSheetCbb, False)).grid(row=rowNum, column=2,
                                                                                            sticky=mSticky,
                                                                                            pady=4)

        rowNum += 1
        Label(master=root, text="文件2第一行：", font=mFont, relief=mRlief).grid(row=rowNum, column=0, sticky=mSticky,
                                                                                pady=4)
        twoRowCbb = Combobox(master=root, textvariable=mTwoRow, font=mFont)
        twoRowCbb.grid(row=rowNum, column=1, sticky=mSticky, pady=4)
        Button(master=root, text="获取", font=mFont, relief=mRlief, command=lambda: self.
               adbSelectRowsName(selectPathTwoTv, twoSheetCbb, twoRowCbb, False)).grid(row=rowNum, column=2,
                                                                                       sticky=mSticky, pady=4)

        rowNum += 1
        Label(master=root, text="新的目录：", font=mFont, relief=mRlief).grid(row=rowNum, column=0, sticky=mSticky,
                                                                             pady=4)
        savedirectoryTv = Label(master=root, text="请选择新的目录", font=mFont, relief=mRlief, wraplength=mWraplength)
        savedirectoryTv.grid(row=rowNum, column=1, sticky=mSticky, pady=4)
        Button(master=root, text="选择", font=mFont, relief=mRlief, command=lambda: self.
               adbNewPathdirectory(savedirectoryTv)).grid(row=rowNum, column=2, sticky=mSticky, pady=4)
        rowNum += 1
        mTypeVar = IntVar(master=root)
        mTypeVar.set(0)
        Label(master=root, text="选择方式", font=mFont, relief=mRlief).grid(row=rowNum, column=0, sticky=mSticky,
                                                                            pady=4)
        Radiobutton(master=root, text="新的文件", value=0, variable=mTypeVar, font=mFont, relief=mRlief).grid(
            row=rowNum, column=1,
            sticky=mSticky, pady=4)
        Radiobutton(master=root, text="已有文件", value=1, variable=mTypeVar, font=mFont, relief=mRlief).grid(
            row=rowNum, column=2,
            sticky=mSticky, pady=4)

        rowNum += 1
        Label(master=root, text="重复文件：", font=mFont, relief=mRlief).grid(row=rowNum, column=0, sticky=mSticky,
                                                                             pady=4)
        saveNewPathTv = Label(master=root, text="重复文件目录", font=mFont, relief=mRlief, wraplength=mWraplength)
        saveNewPathTv.grid(row=rowNum, column=1, sticky=mSticky, pady=4)
        Button(master=root, text="选择", font=mFont, relief=mRlief, command=lambda: self.
               adbNewPath(saveNewPathTv)).grid(row=rowNum, column=2, sticky=mSticky, pady=4)

        rowNum += 1
        Label(master=root, text="重复目录：", font=mFont, relief=mRlief).grid(row=rowNum, column=0, sticky=mSticky,
                                                                             pady=4)
        saveName = Entry(master=root, font=mFont, relief=mRlief)
        saveName.grid(row=rowNum, column=1, sticky=mSticky, pady=4)
        Label(master=root, text=".xlsx", font=mFont, relief=mRlief, ).grid(row=rowNum, column=2, sticky=mSticky, pady=4)

        rowNum += 1
        Label(master=root, text="已有文件：", font=mFont, relief=mRlief).grid(row=rowNum, column=0, sticky=mSticky,
                                                                             pady=4)
        savedirectoryNewTv = Label(master=root, text="请选择已有的文件", font=mFont, relief=mRlief,
                                   wraplength=mWraplength)
        savedirectoryNewTv.grid(row=rowNum, column=1, sticky=mSticky, pady=4)
        Button(master=root, text="选择", font=mFont, relief=mRlief, command=lambda: self.
               adbNewFileData(savedirectoryNewTv)).grid(row=rowNum, column=2, sticky=mSticky, pady=4)
        rowNum += 1
        Button(master=root, text="查重并追加文件1", font=mFont,
               command=lambda: self.adbSave(selectPathOneTv, oneSheetCbb, oneRowCbb, selectPathTwoTv, twoSheetCbb,
                                            twoRowCbb, saveNewPathTv, saveName, mIntVar, savedirectoryTv, mTypeVar,
                                            savedirectoryNewTv)).grid(row=rowNum, column=1,
                                                                      sticky=mSticky, pady=4)

        self.registLisetener(root, mainRoot)
        mainloop()
        pass

    def adbSelectFileOne(self, selectPathTv, mSheet, mRow, isOne: bool, mIntvar):
        if not isOne:
            get = mIntvar.get()
            if get==1:
                self.showWaring("请选择文件2")
                return

        type = {("XLSX", ".xlsx")}
        selectOnePath = filedialog.askopenfilename(filetypes=type)
        oldPath = selectPathTv.cget("text")
        if oldPath != selectOnePath:
            mSheet.set("")
            mRow.set("")
        if selectOnePath is None or selectOnePath == "":
            if isOne:
                selectPathTv.config(text="One文件目录")
            else:
                selectPathTv.config(text="Two文件目录")
            return
        print(f"{selectOnePath}")
        selectPathTv.config(text=selectOnePath)
        pass

    def adbSelectSheeet(self, str: StringVar, selectPathOneTv, oneSheetCbb, isOne: bool):
        selectPath = selectPathOneTv.cget("text")
        typeName = "One文件目录"
        typeNameTwo = "请选择One文件目录"
        if isOne:
            typeName = "One文件目录"
            typeNameTwo = "请选择One文件目录"
        else:
            typeName = "Two文件目录"
            typeNameTwo = "请选择Two文件目录"
        if selectPath is None or selectPath == typeName:
            self.showWaring(typeNameTwo)
            return
        wb: Workbook = None
        wb = load_workbook(selectPath)
        sheetnames = wb.sheetnames
        oneSheetCbb["value"] = sheetnames
        oneSheetCbb.current(0)
        str.set("")
        wb.close()

    pass

    def adbSelectRowsName(self, selectPathTv, comSheet: Combobox, comRow: Combobox, isOne: bool):
        selectPath = selectPathTv.cget("text")
        current = comSheet.current()
        typeParam1 = "One文件目录"
        typeParamWaring1 = "请选择One文件目录"
        if isOne:
            typeParam1 = "One文件目录"
            typeParamWaring1 = "请选择One文件目录"
        else:
            typeParam1 = "Two文件目录"
            typeParamWaring1 = "请选择Two文件目录"

        if selectPath is None or selectPath == typeParam1:
            self.showWaring(typeParamWaring1)
            return
        if current == -1:
            self.showWaring("请选择表名")
            return
        print(f"adbOneRowsName=={selectPath}")
        wb: Workbook = None
        wb = load_workbook(selectPath)
        sheetnames = wb.sheetnames
        selectSheet = sheetnames[current]
        sh = wb[selectSheet]
        # 总共几行
        # rowNumber = sh.max_column
        rows = sh.rows
        # print(f"rowNumebrs==={rowNumber}")
        # if rowNumber==1:
        #     self.showWaring(f"选择表{selectSheet},里面没有内容")
        #     return
        # 共几列
        columns = sh.max_column
        # 第一列数据
        rowList = []
        isExit = False

        for item in rows:
            if item is None or len(item) == 0:
                comRow["value"] = []
                return
            if isExit: break
            childrenIndex = 1
            for children in item:
                rowList.append(children.value)
                print(f"row{children.value}")
                if childrenIndex == columns:
                    isExit = True
                childrenIndex += 1
        if rowList.__len__() == 0:
            self.showWaring("获取表格数据为空")
            return
        comRow["value"] = rowList
        comRow.current(0)
        wb.close()

    pass

    def adbNewPath(self, saveNewPathTv):
        savePath = filedialog.askdirectory()

        if savePath is None or savePath == "":
            saveNewPathTv.config(text="重复文件目录")
            return
        saveNewPathTv.config(text=savePath)

        pass

    def adbSave(self, selectPathOneTv, oneSheetCbb, oneRowCbb, selectPathTwoTv, twoSheetCbb, twoRowCbb, saveNewPathTv,
                saveName, mIntVar, savedirectoryTv, mTypeVar, savedirectoryNewTv):
        selectOnePath = self.getLabelValue(selectPathOneTv)
        oneSheet = oneSheetCbb.current()
        oneRow = oneRowCbb.current()
        selectTwoPath = self.getLabelValue(selectPathTwoTv)
        twoSheel = twoSheetCbb.current()
        twoRow = twoRowCbb.current()
        saveNewPath = self.getLabelValue(saveNewPathTv)
        saveName = self.getEntryValue(saveName)

        selecetDisatory = self.getLabelValue(savedirectoryTv)
        mSelectFielOrDistory = mIntVar.get()

        selectRepeatNew = self.getLabelValue(savedirectoryNewTv)
        mSelectRepeat = mTypeVar.get()

        if selectOnePath is None or selectOnePath == "One文件目录":
            self.showWaring("请选择One文件目录")
            return
        if oneSheet == -1:
            self.showWaring("请选择One表名")
            return
        if oneRow == -1:
            self.showWaring("请选择One第一行数据")
            return
        if selectTwoPath is None or selectTwoPath == "One文件目录":
            self.showWaring("请选择One文件目录")
            return
        if twoSheel == -1:
            self.showWaring("请选择One表名")
            return
        if twoRow == -1:
            self.showWaring("请选择One第一行数据")
            return

        wb: Workbook = None
        # 第一行数据
        wb = load_workbook(selectOnePath)
        mOnesheetnames = wb.sheetnames
        sheetOne = mOnesheetnames[oneSheet]
        shOne = wb[sheetOne]
        rowsOne = shOne.rows

        print(f"crun=={oneRow}")
        # itemOneIndex=0
        mOneGETlist = []
        for itemOne in rowsOne:
            # itemOneIndex+=1
            # print(f'itemOneIndex=={itemOne},{itemOne}')
            childIndex = 0
            for childOne in itemOne:
                if childIndex == oneRow:
                    mOneGETlist.append(childOne.value)
                childIndex += 1
                print(f"chileOn={childOne.value},childIndex={childIndex}")
        wb.close()
        print(f"OnEList获取数据=={mOneGETlist}")
        if mSelectFielOrDistory == 0:
            wb = load_workbook(selectTwoPath)
            sheetnamesTwo = wb.sheetnames
            sheetTwo = sheetnamesTwo[twoSheel]
            shTwo = wb[sheetTwo]
            rowsTwo = shTwo.rows
            twoGetlist = []
            for itemTwo in rowsTwo:
                childIndex = 0
                for childtTwo in itemTwo:
                    if childIndex == twoRow:
                        twoGetlist.append(childtTwo.value)
                    childIndex += 1

            wb.close()

            print(f"TwoEList获取数据=={twoGetlist}")
            newGetlist = []
            newGetlist.extend(mOneGETlist)
            repead = []
            repeadNo = []
            for item in twoGetlist:
                if item in newGetlist:
                    repead.append(item)
                else:
                    repeadNo.append(item)
                    newGetlist.append(item)
        else:  # 新目录
            if selecetDisatory == "请选择新的目录":
                self.showWaring("请选择新的目录")
                return
            newDirectory = os.listdir(selecetDisatory)

            if newDirectory.__len__() == 0:
                self.showWaring("新的目录下未有文件")
                return
            newGetlist = []
            newGetlist.extend(mOneGETlist)
            repead = []
            repeadNo = []
            for item in newDirectory:
                if item in newGetlist:
                    repead.append(item)
                else:
                    repeadNo.append(item)
                    newGetlist.append(item)
        # newGetlist.extend(twoGetlist)
        # 判断是否存在重复数据，提示保存
        if repead.__len__() > 0 and repeadNo.__len__() > 0:
            askyesno = messagebox.askyesno("提示", "是否要保持重复数据")
            print(f"===asno=={askyesno}")
            if askyesno:
                if mSelectRepeat == 0:
                    if saveNewPath is None or saveNewPath == "重复文件目录":
                        self.showWaring("请选择重复文件目录")
                        return
                    if saveName is None or saveName == "":
                        self.showWaring("请输入新的表名")
                        return
                    saveOver = saveNewPath + "/" + saveName + ".xlsx"
                    if os.path.exists(saveOver):
                        # self.showWaring(f"保存的文件{saveName}.xlsx 已经存在")
                        print(f"重复保存的表=={saveOver}")

                        self.saveExcleOne(repeadNo=repead, selectOnePath=saveOver, oneSheet=saveName,
                                          selectTwoPath=selectTwoPath, isSheetNameStr=True)
                        return
                else:
                    if selectRepeatNew == "请选择已有的文件":
                        self.showWaring("请选择已有的文件")
                        return
                    saveOver = selectRepeatNew
                    self.saveExcleOne(repeadNo=repead, selectOnePath=selectRepeatNew, oneSheet=saveName,
                                      selectTwoPath=selectRepeatNew, isSheetNameStr=True)

                wb = Workbook()
                sh = wb.create_sheet(saveName, 0)
                sh.cell(row=1, column=1).value = f"{datetime.datetime.now()}=={selectTwoPath}重复数据"
                for postion in range(len(repead)):
                    sh.cell(row=postion + 2, column=1).value = repead[postion]
                try:
                    wb.save(saveOver)
                    wb.close()
                except:
                    self.showWaring(f"保存的文件{saveOver},正在编辑或者异常")
                    return
                self.saveExcleOne(repeadNo, selectOnePath, oneSheet, selectTwoPath, False)
                # if self.showAskQuestion("重复数据保存成功，是否跳转对应目录"):
                #     os.startfile(saveNewPath)
            else:
                self.saveExcleOne(repeadNo, selectOnePath, oneSheet, selectTwoPath, False)
            # 执行追加数据
        else:
            self.saveExcleOne(repeadNo, selectOnePath, oneSheet, selectTwoPath, False)

    def saveExcleOne(self, repeadNo, selectOnePath, oneSheet, selectTwoPath, isSheetNameStr: bool):
        if repeadNo.__len__() == 0:
            messagebox.showerror("提示", "无需要保持的数据，合并数据都已存在")
            return

        wb: Workbook = None
        # 第一行数据
        wb = load_workbook(selectOnePath)
        mOnesheetnames = wb.sheetnames
        if isSheetNameStr:
            sheetOne = oneSheet
        else:
            sheetOne = mOnesheetnames[oneSheet]
        print(f"当前表名=={sheetOne}")
        sh = wb[sheetOne]
        rowsOne = sh.max_row + 1
        sh.cell(row=rowsOne, column=1).value = f"{datetime.datetime.now()}=={selectTwoPath}追加数据"
        for index in range(len(repeadNo)):
            sh.cell(row=rowsOne + index + 1, column=1).value = repeadNo[index]
        try:
            wb.save(selectOnePath)
            wb.close()
        except Exception as e:
            print(e)
            self.showWaring(f"保存的文件{selectOnePath},正在编辑或者异常")
            return
        messagebox.showinfo("成功", '追加成功')
        pass

    def adbNewPathdirectory(self, savedirectoryTv):
        self.selectFileDialog(savedirectoryTv, FileDialogType.DIRECTORY, "", "请选择新的目录", True)
        pass

    def adbNewFileData(self, savedirectoryNewTv):
        fileType = [("Xlsx", ".xlsx")]
        dialog = filedialog.askopenfile(filetypes=fileType)
        if dialog is not None:
            path = dialog.name
        else:
            path = "请选择已有的文件"
        savedirectoryNewTv.config(text=f"{path}")
        # self.selectFileDialog(savedirectoryNewTv,FileDialogType.DIRECTORY,"","请选择新的文件",True)
        pass
