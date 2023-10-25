#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2023/7/19 21:16
# @Author  : backpacker
# @File    : MergeXlsx.py
# @Description : $合并xlsx
import os
from tkinter import *
from tkinter.ttk import Combobox

from Base.MainQuit import MainQuit
from tkinter import *
from tkinter import ttk
from tkinter import filedialog
from openpyxl import Workbook,load_workbook


class MergeXlsx(MainQuit):
    def showDialog(self, mainRoot):
        root = Tk()
        root.title("查重工具")
        root.config(bg="light gray")
        root.geometry("640x540")  # 设置窗口大小
        mFont = 10
        mSticky = W + E + S + N
        mWraplength = 300,
        mRlief = "groove"
        rowNum = 0
        mOneSheet = StringVar(master=root)
        mOneRow = StringVar(master=root)
        mTwoSheet = StringVar(master=root)
        mTwoRow = StringVar(master=root)
        # ==============================================================
        Label(master=root, text="原文件One：", font=mFont, relief=mRlief).grid(row=rowNum, column=0, sticky=mSticky,
                                                                              pady=4)
        selectPathOneTv = Label(master=root, text="One文件目录", wraplength=mWraplength, font=mFont, relief=mRlief)
        selectPathOneTv.grid(row=rowNum, column=1, sticky=mSticky, pady=4)
        Button(master=root, text="选择", font=mFont, relief=mRlief, command=lambda: self.
               adbSelectFileOne(selectPathOneTv, mOneSheet, mOneRow,True)).grid(row=rowNum, column=2,
                                                                                            sticky=mSticky, pady=4)
        len__ = root.children.__len__()
        for index in range(len__):
            root.columnconfigure(index, weight=1)
        # ==============================================================
        rowNum += 1
        Label(master=root, text="文件One表名：", font=mFont, relief=mRlief).grid(row=rowNum, column=0, sticky=mSticky,
                                                                                pady=4)
        oneSheetCbb = Combobox(master=root, textvariable=mOneSheet, font=mFont,postcommand=lambda :self.adbComboChange(mOneRow))
        oneSheetCbb.grid(row=rowNum, column=1, sticky=mSticky, pady=4)
        Button(master=root, text="获取", font=mFont, relief=mRlief, command=lambda: self.
               adbSelectSheeet(mOneRow, selectPathOneTv, oneSheetCbb, True)).grid(row=rowNum, column=2, sticky=mSticky, pady=4)

        rowNum += 1
        Label(master=root, text="文件One第一行：", font=mFont, relief=mRlief).grid(row=rowNum, column=0, sticky=mSticky,
                                                                                pady=4)
        oneRowCbb = Combobox(master=root, textvariable=mOneRow, font=mFont)
        oneRowCbb.grid(row=rowNum, column=1, sticky=mSticky, pady=4)
        Button(master=root, text="获取", font=mFont, relief=mRlief, command=lambda: self.
               adbSelectRowsName(selectPathOneTv, oneSheetCbb, oneRowCbb, True)).grid(row=rowNum, column=2, sticky=mSticky, pady=4)

        rowNum += 1
        Label(master=root, text="原文件Two：", font=mFont, relief=mRlief).grid(row=rowNum, column=0, sticky=mSticky,
                                                                              pady=4)
        selectPathTwoTv = Label(master=root, text="Two文件目录", font=mFont, relief=mRlief, wraplength=mWraplength)
        selectPathTwoTv.grid(row=rowNum, column=1, sticky=mSticky, pady=4)
        Button(master=root, text="选择", font=mFont, relief=mRlief, command=lambda: self.
               adbSelectFileOne(selectPathTwoTv,mTwoSheet,mTwoRow,False)).grid(row=rowNum, column=2, sticky=mSticky, pady=4)

        rowNum += 1
        Label(master=root, text="文件Two表名：", font=mFont, relief=mRlief).grid(row=rowNum, column=0, sticky=mSticky,
                                                                                pady=4)
        twoSheetCbb = Combobox(master=root, textvariable=mTwoSheet, font=mFont,postcommand=lambda :self.adbComboChange(mTwoRow))
        twoSheetCbb.grid(row=rowNum, column=1, sticky=mSticky, pady=4)
        Button(master=root, text="获取", font=mFont, relief=mRlief, command=lambda: self.
               adbSelectSheeet(mTwoRow, selectPathTwoTv, twoSheetCbb, False)).grid(row=rowNum, column=2, sticky=mSticky, pady=4)

        rowNum += 1
        Label(master=root, text="文件Two第一行：", font=mFont, relief=mRlief).grid(row=rowNum, column=0, sticky=mSticky,
                                                                                pady=4)
        twoRowCbb = Combobox(master=root, textvariable=mTwoRow, font=mFont)
        twoRowCbb.grid(row=rowNum, column=1, sticky=mSticky, pady=4)
        Button(master=root, text="获取", font=mFont, relief=mRlief, command=lambda: self.
               adbSelectRowsName(selectPathTwoTv, twoSheetCbb, twoRowCbb, False)).grid(row=rowNum, column=2, sticky=mSticky, pady=4)

        rowNum += 1
        Label(master=root, text="合并文件：", font=mFont, relief=mRlief).grid(row=rowNum, column=0, sticky=mSticky,
                                                                             pady=4)
        saveNewPathTv = Label(master=root, text="新文件目录", font=mFont, relief=mRlief, wraplength=mWraplength)
        saveNewPathTv.grid(row=rowNum, column=1, sticky=mSticky, pady=4)
        Button(master=root, text="选择", font=mFont, relief=mRlief, command=lambda: self.
               adbNewPath(saveNewPathTv)).grid(row=rowNum, column=2, sticky=mSticky, pady=4)

        rowNum += 1
        Label(master=root, text="合并文件名称：", font=mFont, relief=mRlief).grid(row=rowNum, column=0, sticky=mSticky,
                                                                                 pady=4)
        saveName = Entry(master=root, font=mFont, relief=mRlief)
        saveName.grid(row=rowNum, column=1, sticky=mSticky, pady=4)
        Label(master=root, text=".xlsx", font=mFont, relief=mRlief, ).grid(row=rowNum, column=2, sticky=mSticky, pady=4)

        rowNum += 1
        Button(master=root, text="确认合并", font=mFont, command=lambda: self.adbSave(selectPathOneTv,oneSheetCbb,oneRowCbb,selectPathTwoTv,twoSheetCbb,twoRowCbb,saveNewPathTv,saveName)).grid(row=rowNum, column=1,
                                                                                         sticky=mSticky, pady=4)

        self.registLisetener(root, mainRoot)
        mainloop()
        pass

    pass

    def adbSelectFileOne(self, selectPathTv, mSheet, mRow, isOne:bool):
        type = {("XLSX", ".xlsx")}
        selectOnePath = filedialog.askopenfilename(filetypes=type)
        oldPath= selectPathTv.cget("text")
        if oldPath != selectOnePath:
            mSheet.set("")
            mRow.set("")
        if selectOnePath is None or selectOnePath == "":
            if isOne:
                selectPathTv.config(text="One文件目录")
            else:
                selectPathTv.config(text="Two文件目录")
            return
        print(f"{selectOnePath}")
        selectPathTv.config(text=selectOnePath)

        pass

    def adbSelectSheeet(self, str:StringVar, selectPathOneTv, oneSheetCbb, isOne:bool):
        selectPath = selectPathOneTv.cget("text")

        typeName="One文件目录"
        typeNameTwo="请选择One文件目录"
        if isOne:
            typeName = "One文件目录"
            typeNameTwo = "请选择One文件目录"
        else:
            typeName="Two文件目录"
            typeNameTwo = "请选择Two文件目录"

        if selectPath is None or selectPath==typeName:
            self.showWaring(typeNameTwo)
            return
        wb:Workbook=None
        wb = load_workbook(selectPath)
        sheetnames = wb.sheetnames
        oneSheetCbb["value"]=sheetnames
        oneSheetCbb.current(0)
        str.set("")
        wb.close()
    pass

    def adbSelectRowsName(self, selectPathTv, comSheet:Combobox, comRow:Combobox, isOne:bool):
        selectPath = selectPathTv.cget("text")
        current = comSheet.current()

        typeParam1="One文件目录"
        typeParamWaring1="请选择One文件目录"

        if isOne:
            typeParam1="One文件目录"
            typeParamWaring1 = "请选择One文件目录"
        else:
            typeParam1 = "Two文件目录"
            typeParamWaring1 = "请选择Two文件目录"

        if selectPath is None or selectPath==typeParam1:
            self.showWaring(typeParamWaring1)
            return
        if current==-1:
            self.showWaring("请选择表名")
            return
        print(f"adbOneRowsName=={selectPath}")
        wb:Workbook=None
        wb=load_workbook(selectPath)
        sheetnames = wb.sheetnames
        selectSheet = sheetnames[current]
        sh = wb[selectSheet]
        #总共几行
        # rowNumber = sh.max_column
        rows = sh.rows
        # print(f"rowNumebrs==={rowNumber}")
        # if rowNumber==1:
        #     self.showWaring(f"选择表{selectSheet},里面没有内容")
        #     return
        #共几列
        columns = sh.max_column
        #第一列数据
        rowList=[]
        isExit=False

        for item in rows:
            if item is None or len(item)==0:
                comRow["value"]=[]
                return
            if isExit: break
            childrenIndex = 1
            for children in item:
                rowList.append(children.value)
                print(f"row{children.value}")
                if childrenIndex==columns:
                    isExit=True
                childrenIndex+=1
        if rowList.__len__()==0:
            self.showWaring("获取表格数据为空")
            return
        comRow["value"]=rowList
        comRow.current(0)
        wb.close()

    pass

    def adbNewPath(self, saveNewPathTv):
        savePath = filedialog.askdirectory()

        if savePath is None or savePath=="":
            saveNewPathTv.config(text="新文件目录")
            return
        saveNewPathTv.config(text=savePath)

        pass

    def adbSave(self, selectPathOneTv, oneSheetCbb, oneRowCbb, selectPathTwoTv, twoSheetCbb, twoRowCbb, saveNewPathTv,
                saveName):
        selectOnePath = self.getLabelValue(selectPathOneTv)
        oneSheet = oneSheetCbb.current()
        oneRow = oneRowCbb.current()
        selectTwoPath = self.getLabelValue(selectPathTwoTv)
        twoSheel = twoSheetCbb.current()
        twoRow = twoRowCbb.current()
        saveNewPath = self.getLabelValue(saveNewPathTv)
        saveName = self.getEntryValue(saveName)


        if selectOnePath is None or selectOnePath=="One文件目录":
            self.showWaring("请选择One文件目录")
            return
        if oneSheet ==-1:
            self.showWaring("请选择One表名")
            return
        if oneRow==-1:
            self.showWaring("请选择One第一行数据")
            return
        if selectTwoPath is None or selectTwoPath=="One文件目录":
            self.showWaring("请选择One文件目录")
            return
        if twoSheel ==-1:
            self.showWaring("请选择One表名")
            return
        if twoRow==-1:
            self.showWaring("请选择One第一行数据")
            return

        if saveNewPath is None or saveNewPath=="新文件目录":
            self.showWaring("请选择新文件目录")
            return

        if saveName is None or saveName=="":
            self.showWaring("请输入新的表名")
            return

        wb:Workbook=None
        #第一行数据
        wb=load_workbook(selectOnePath)
        mOnesheetnames = wb.sheetnames
        sheetOne = mOnesheetnames[oneSheet]
        shOne=wb[sheetOne]
        rowsOne = shOne.rows

        print(f"crun=={oneRow}")
        # itemOneIndex=0
        mOneGETlist=[]
        for itemOne in rowsOne:
            # itemOneIndex+=1
            # print(f'itemOneIndex=={itemOne},{itemOne}')
            childIndex=0
            for childOne in itemOne:
                if childIndex==oneRow:
                    mOneGETlist.append(childOne.value)
                childIndex+=1
                print(f"chileOn={childOne.value},childIndex={childIndex}")
        wb.close()

        print(f"OnEList获取数据=={mOneGETlist}")

        wb=load_workbook(selectTwoPath)
        sheetnamesTwo = wb.sheetnames
        sheetTwo = sheetnamesTwo[twoSheel]
        shTwo = wb[sheetTwo]
        rowsTwo = shTwo.rows
        twoGetlist=[]
        for itemTwo in rowsTwo:
            childIndex=0
            for childtTwo in itemTwo:
                if childIndex==twoRow:
                    twoGetlist.append(childtTwo.value)
                childIndex += 1

        wb.close()

        print(f"TwoEList获取数据=={twoGetlist}")
        newGetlist=[]
        newGetlist.extend(mOneGETlist)
        newGetlist.extend(twoGetlist)

        saveOver = saveNewPath+"/" + saveName+".xlsx"

        if os.path.exists(saveOver):
            self.showWaring(f"保存的文件{saveName}.xlsx 已经存在")
            return

        wb=Workbook()
        sh=wb.create_sheet(saveName)
        if newGetlist.__len__()==0:
            self.showWaring("要保存的数据为空")
            return

        for postion in range(len(newGetlist)):
            sh.cell(row=postion+1,column=1).value=newGetlist[postion]
        try:
            wb.save(saveOver)
            wb.close()
        except:
            self.showWaring(f"保存的文件{saveOver},正在编辑或者异常")
            return
        if self.showAskQuestion("保存成功，是否跳转对应目录"):
            os.startfile(saveNewPath)
        pass


