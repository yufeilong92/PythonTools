#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2023/10/20 23:02
# @Author  : backpacker
# @File    : addOneExcel.py
# @Description : $
import os
from base64 import encode, decode
from tkinter import Tk, filedialog, messagebox
from tkinter import *

from openpyxl import Workbook ,load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from Base.FileDialogType import FileDialogType
from Base.MainQuit import MainQuit
import win32clipboard

from Base.TypeBgColor import TypeBgColor


class AddOneExecle(MainQuit):

    def adbSelectFileOne(self, selectPathTv):
        type = {("XLSX", ".xlsx")}
        selectOnePath = filedialog.askopenfilename(filetypes=type)
        oldPath = selectPathTv.cget("text")
        if oldPath == selectOnePath:
            return
        if selectOnePath is None or selectOnePath == "":
            selectPathTv.config(text="excel的地址")
            return
        print(f"{selectOnePath}")
        selectPathTv.config(text=selectOnePath)

        pass

    def showDialog(self, mainRoot):
        root = Tk()
        root.title("保存数据到exle")
        root.config(bg="light gray")
        root.geometry("640x600")  # 设置窗口大小
        mFont = 10
        mSticky = W + E + S + N
        mWraplength = 300,
        mRlief = "groove"
        rowNum=0
        showContext:Label
        Label(master=root, text="现有Excel:", font=mFont, relief=mRlief).grid(row=rowNum, column=0, sticky=mSticky,
                                                                                  pady=4)
        selectExcelTV = Label(master=root, text="excel的地址", font=mFont,
                            wraplength=mWraplength, relief=mRlief)
        selectExcelTV.grid(row=rowNum, column=1, sticky=mSticky, pady=4)

        Button(master=root, text="选择", font=mFont, relief=mRlief,
               command=lambda: self.adbSelectFileOne(selectExcelTV)).grid(row=rowNum, column=2, sticky=mSticky,
                                                                                      pady=4)
        len__ = root.children.__len__()
        for index in range(len__):
            root.columnconfigure(index, weight=1)
        # ===============标题权重===================
        rowNum += 1
        Label(master=root, text="创建新的:", font=mFont, relief=mRlief).grid(row=rowNum, column=0, sticky=mSticky,
                                                                              pady=4)
        createExcelFileTV = Label(master=root, text="excel的目录", font=mFont,
                              wraplength=mWraplength, relief=mRlief)
        createExcelFileTV.grid(row=rowNum, column=1, sticky=mSticky, pady=4)

        Button(master=root, text="选择", font=mFont, relief=mRlief,
               command=lambda: self.selectFileDialog(createExcelFileTV,FileDialogType.DIRECTORY,"excel的目录", "",
                                                     False)).grid(row=rowNum, column=2, sticky=mSticky,
                                                                          pady=4)
        rowNum += 1
        Label(master=root, text="新的名称:", font=mFont, relief=mRlief).grid(row=rowNum, column=0, sticky=mSticky,
                                                                             pady=4)
        createNameEt = Entry(master=root, text="请输入名称", font=mFont)
        createNameEt.grid(row=rowNum, column=1, sticky=mSticky, pady=4)

        Label(master=root, text=".xlsx", font=mFont, relief=mRlief).grid(row=rowNum, column=2, sticky=mSticky,

                                                                          pady=4)
        rowNum += 1
        Label(master=root, text="保存列One:", font=mFont, relief=mRlief).grid(row=rowNum, column=0, sticky=mSticky, pady=4)
        createNameOneEt = Entry(master=root,  font=mFont)
        createNameOneEt.grid(row=rowNum, column=1,columnspan=2, sticky=mSticky, pady=4)


        rowNum += 1
        Label(master=root, text="保存列Two:", font=mFont, relief=mRlief).grid(row=rowNum, column=0, sticky=mSticky,
                                                                             pady=4)
        createNameTwoEt=Entry(master=root,font=mFont)
        createNameTwoEt.grid(row=rowNum, column=1,columnspan=2, sticky=mSticky, pady=4)


        rowNum += 1
        Button(master=root, text="清空列One和Two", font=mFont, relief=mRlief,
               command=lambda: self.clearOneTwo(createNameOneEt,createNameTwoEt,showContext)).grid(row=rowNum, column=0,columnspan=3, sticky=mSticky,
                                                                  pady=4)

        rowNum += 1
        Button(master=root, text="查询", font=mFont, relief=mRlief,
               command=lambda: self.queryList(selectExcelTV,createExcelFileTV,createNameEt,createNameOneEt,createNameTwoEt,showContext)).grid(row=rowNum, column=0,columnspan=3, sticky=mSticky,
                                                                  pady=4)
        rowNum += 1
        Button(master=root, text="添加", font=mFont, relief=mRlief,
               command=lambda: self.addSaveExcle(selectExcelTV,createExcelFileTV,createNameEt,createNameOneEt,createNameTwoEt,showContext)).grid(row=rowNum, column=0,columnspan=3, sticky=mSticky,
                                                                  pady=4)
        rowNum += 1
        Button(master=root, text="粘贴One", font=mFont, relief=mRlief,
               command=lambda: self.postadd(selectExcelTV,createExcelFileTV,createNameEt,createNameOneEt,createNameTwoEt,showContext)).grid(row=rowNum, column=0, sticky=mSticky,
                                                                  pady=4)
        Button(master=root, text="粘贴Two", font=mFont, relief=mRlief,
               command=lambda: self.postaddTwo(selectExcelTV, createExcelFileTV, createNameEt, createNameOneEt,
                                            createNameTwoEt, showContext)).grid(row=rowNum, column=2,
                                                                                sticky=mSticky,
                                                                                pady=4)
        rowNum += 1
        Label(master=root,text="提示：默认优选添加到现有excele,默认表一,,其次是新建excele",font=mFont).grid(row=rowNum,column=0,columnspan=3,sticky=mSticky,pady=4)

        rowNum += 1
        showContext=Label(master=root,height=10,font=mFont,wraplength=600,bg="#5EA4DE")
        showContext.grid(row=rowNum,column=0,columnspan=3,sticky=mSticky,pady=4)

        self.registLisetener(root, mainRoot)
        mainloop()
        pass

    pass



    def clearOneTwo(self, createNameOneEt, createNameTwoEt,showContext):
        createNameTwoEt.delete(0,"end")
        createNameOneEt.delete(0,"end")
        self.setTvContext(showContext,"",TypeBgColor.info)
        pass
    #selectExcelTV 现有的excel
    #createExcelFileTV 创建的excle目录
    #createNameEt 创建的名称
    #createNameOneEt one 数据
    #createNameTwoEt two 数据
    def addSaveExcle(self, selectExcelTV, createExcelFileTV, createNameEt, createNameOneEt, createNameTwoEt,showContext):
        #现有的数据

        try:
            selecetExcele = selectExcelTV.cget("text")
            createPath = createExcelFileTV.cget("text")
            createSheetName = createNameEt.get()
            savePath=None
            wb:Workbook=None
            sh:Worksheet=None
            isrep=False;


            if selecetExcele == "excel的地址":
                if createPath == "" or createSheetName == "":
                    self.setTvContext(showContext, "温馨提示\n   请选择要保存的Excele", TypeBgColor.waring)
                    return
            if selecetExcele!="excel的地址":
                savePath=selecetExcele
                wb=load_workbook(selecetExcele)
                mOnesheetnames = wb.sheetnames
                print(f"{mOnesheetnames}")
                sheetOne = mOnesheetnames[0]
                sh = wb[sheetOne]
                print(f"{mOnesheetnames}//{sh}//{sh.max_row}")
                rowsOne = sh.max_row+1

                isrep=True
                pass


            else:
            #创建新表添加数据
                savePath=createPath+"/"+createSheetName+".xlsx"
                if os.path.exists(savePath):
                    self.setTvContext(showContext,f"温馨提示\n  {savePath},文件已经存在",TypeBgColor.waring)
                    #messagebox.showwarning("温馨提示",f"{savePath},文件已经存在")
                    return

                wb = Workbook()
                sh = wb.create_sheet(createSheetName,0)
                rowsOne=1
                isrep = True
                pass
            oneData = createNameOneEt.get()
            if oneData is None or oneData =="":
                self.setTvContext(showContext, f"温馨提示\n  列One 数据为空",TypeBgColor.waring)
                #messagebox.showwarning("温馨提示","列One 数据为空")
                return
            twoData = createNameTwoEt.get()
            if twoData is None or twoData=="":
                self.setTvContext(showContext, f"温馨提示\n  列Two 数据为空",TypeBgColor.waring)
                #messagebox.showwarning("温馨提示", "列Two 数据为空")
                return
           #=========判断是否重复============
            if isrep:
                sh_rows = sh.rows
                #maxcolumn = sh.max_column#列
                #maxrow = sh.max_row #行
                #第一列数据
                # oneSheetData = sh.iter_cols(1, 1, values_only=True)
                # mOneList=[]
                # for item in oneSheetData:
                #     if item is not  None:
                #         for itemA in item:
                #             if itemA is not  None:
                #                 mOneList.append(itemA)
                # print(f"oneList=={mOneList}")
                # if mOneList.__len__()!=0 :
                #     if oneData in mOneList:
                #         self.setTvContext(showContext, f"温馨提示\n  数据One{oneData}数据重复")
                #         #messagebox.showwarning("温馨提示",f"数据One{oneData}数据重复")
                #         return
                #
                # # 第一列数据
                # twoSheetData = sh.iter_cols(2, 2, values_only=True)
                # mTwoList = []
                # for item in twoSheetData:
                #     if item is not None:
                #         for itemA in item:
                #             if itemA is not None:
                #                 mTwoList.append(itemA)
                # print(f"twoList=={mTwoList}")
                # if mTwoList.__len__()!=0:
                #     if twoData in mTwoList:
                #         self.setTvContext(showContext, f"温馨提示\n  数据Two{twoData}数据重复")
                #         #messagebox.showwarning("温馨提示", f"数据Two{twoData}数据重复")
                mAlllist = []
                oneSheetData = sh.iter_cols(1, 1, values_only=True)
                for item in oneSheetData:
                    if item is not None:
                        for itemA in item:
                            if itemA is not None:
                                mAlllist.append(itemA)
                # 第2列数据
                twoSheetData = sh.iter_cols(2, 2, values_only=True)
                for item in twoSheetData:
                    if item is not None:
                        for itemA in item:
                            if itemA is not None:
                                mAlllist.append(itemA)
                print(f"mall=={mAlllist}")
                if mAlllist.__len__() == 0:
                    self.setTvContext(showContext,
                                      f"温馨提示\n========Success========\n数据数据one={oneData}\n 数据two={twoData}可以添加",TypeBgColor.Success)
                    return
                if oneData != "" and oneData in mAlllist:
                    self.setTvContext(showContext, f"温馨提示\n========Error========\n数据列One={oneData}数据重复",TypeBgColor.error)
                    return
                if twoData != "" and twoData in mAlllist:
                    self.setTvContext(showContext, f"温馨提示\n========Error========\n数据列two{twoData}数据重复",TypeBgColor.error)
                    return
                # self.setTvContext(showContext, f"温馨提示\n========Success========\n数据{oneData}{twoData}可以添加")
                        #=============保存=============
            sh.cell(row=rowsOne , column=1).value = f"{oneData}"
            sh.cell(row=rowsOne , column=2).value = f"{twoData}"

            wb.save(savePath)
            wb.close()
            self.setTvContext(showContext, f"温馨提示\n =======Success========\n 要保存的文件{savePath}成功",TypeBgColor.AddSuccess)
            #messagebox.showinfo("温馨提示", f"要保存的文件{savePath}成功")
        except:
            self.setTvContext(showContext, f"温馨提示\n 要保存的文件{savePath}正在打开，或者异常，请关闭重试",TypeBgColor.error)
            #messagebox.showerror("温馨提示", f"要保存的文件{savePath}正在打开，或者异常，请关闭重试")


            pass
    def setTvContext(self,showContxt:Label,strContxt:str,color:TypeBgColor):
        showContxt.config(text=strContxt)
        bg="#5EA4DE"
        if color==TypeBgColor.waring:
            bg = "#C4C400"
        elif color==TypeBgColor.info:
            bg="#2894FF"
        elif color==TypeBgColor.error:
            bg="#EA0000"
        elif color==TypeBgColor.Success:
            bg="#00BB00"
        elif color == TypeBgColor.AddSuccess:
            bg = "#CA8EFF"
        elif color==TypeBgColor.defend:
            bg=""
        print(f"{bg}")
        if bg=="" or bg is None:
            return
        showContxt.config(bg=f"{bg}")
        pass

    def queryList(self, selectExcelTV, createExcelFileTV, createNameEt, createNameOneEt, createNameTwoEt, showContext):
        selecetExcele = selectExcelTV.cget("text")
        wb: Workbook = None
        sh: Worksheet = None
        if selecetExcele != "excel的地址":
            wb = load_workbook(selecetExcele)
            mOnesheetnames = wb.sheetnames
            print(f"{mOnesheetnames}")
            sheetOne = mOnesheetnames[0]
            sh = wb[sheetOne]
            print(f"{mOnesheetnames}//{sh}//{sh.max_row}")
            pass
        else:
            self.setTvContext(showContext, "温馨提示\n   请选择要查询的Excele",TypeBgColor.waring)
            return
        oneData = createNameOneEt.get()
        twoData = createNameTwoEt.get()

        if oneData=="" and twoData=="":
            self.setTvContext(showContext,"温馨提示\n   列表One 或者列表two 不能都为空",TypeBgColor.waring)
            return
        # =========判断是否重复============
        sh_rows = sh.rows
        # maxcolumn = sh.max_column#列
        # maxrow = sh.max_row #行
        # 第一列数据
        mAlllist=[]
        oneSheetData = sh.iter_cols(1, 1, values_only=True)
        for item in oneSheetData:
            if item is not None:
                for itemA in item:
                    if itemA is not None:
                        mAlllist.append(itemA)
        # 第2列数据
        twoSheetData = sh.iter_cols(2, 2, values_only=True)
        for item in twoSheetData:
            if item is not None:
                for itemA in item:
                    if itemA is not None:
                        mAlllist.append(itemA)
        print(f"mall=={mAlllist}")
        if mAlllist.__len__()==0:
            self.setTvContext(showContext,f"温馨提示\n========Success========\n数据数据one={oneData}\n 数据two={twoData}可以添加",TypeBgColor.Success)
            return
        if oneData!="" and oneData in mAlllist:
            self.setTvContext(showContext,f"温馨提示\n========Error========\n数据列One={oneData}数据重复",TypeBgColor.error)
            return
        if twoData != "" and twoData in mAlllist:
            self.setTvContext(showContext, f"温馨提示\n========Error========\n数据列two{twoData}数据重复",TypeBgColor.error)
            return
        self.setTvContext(showContext, f"温馨提示\n========Success========\n{oneData}\n{twoData}可以添加",TypeBgColor.Success)
        pass

    def postadd(self, selectExcelTV, createExcelFileTV, createNameEt, createNameOneEt, createNameTwoEt, showContext):
        dataone = createNameOneEt.get()
        if dataone is not None or dataone != "":
            createNameOneEt.delete(0, END)

        datatwo = createNameTwoEt.get()
        if datatwo is not None or datatwo != "":
            createNameTwoEt.delete(0, END)

        win32clipboard.OpenClipboard()
        context=win32clipboard.GetClipboardData(win32clipboard.CF_UNICODETEXT)
        win32clipboard.CloseClipboard()
        print(f"拷贝数one=={context}")

        createNameOneEt.insert(0, context)
        self.setTvContext(showContext, "", TypeBgColor.info)
        pass
    def postaddTwo(self, selectExcelTV, createExcelFileTV, createNameEt, createNameOneEt, createNameTwoEt, showContext):
        cget = createNameTwoEt.get()
        if cget is not None or cget != "":
            createNameTwoEt.delete(0, END)
        win32clipboard.OpenClipboard()
        context=win32clipboard.GetClipboardData(win32clipboard.CF_UNICODETEXT)
        win32clipboard.CloseClipboard()
        print(f"拷贝数two=={context}")

        createNameTwoEt.insert(0,context)
        self.setTvContext(showContext, "", TypeBgColor.info)
        pass